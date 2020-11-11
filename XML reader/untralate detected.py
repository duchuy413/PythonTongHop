import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

print("input source file name: ")
inputFileName = input()


print("input translated file name: ")
translatedFileName = input()

wb = Workbook()
dest_filename = 'result.xlsx'
sheet = wb.active

sheet.cell(column=1, row=1, value="key")
sheet.cell(column=2, row=1, value="eng")
sheet.cell(column=3, row=1, value="translated")


tree = ET.parse(inputFileName)
root = tree.getroot()
print(root)

i = 1
for child in root:
    sheet.cell(column=1, row=i+1, value=child.attrib["name"])
    i += 1

i = 1
for child in root:
    sheet.cell(column=2, row=i+1, value=child.text)
    i += 1

dic = {}

newroot = ET.parse(translatedFileName).getroot()

for child in newroot:
    dic[child.attrib["name"]] = child.text

untranslated = []

i = 1
for child in root:
    if child.attrib["name"] in dic:
        sheet.cell(column=3, row=i+1, value=dic[child.attrib["name"]])
    else:
        sheet.cell(column=3, row=i+1, value="NOT TRANSLATED")
        print ("Line " + str(i) + " values: " + child.attrib["name"])
        untranslated.append(child.attrib["name"])
    i += 1

wb.save(filename = dest_filename)














