import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

print("input source file name: ")
inputFileName = input()


print("input translated file name: ")
translatedFileName = input()

wb = Workbook()
dest_filename = 'result.xlsx'
sheet = wb.active

sheet.cell(column=1, row=1, value="key")
sheet.cell(column=2, row=1, value="eng")
sheet.cell(column=3, row=1, value="translated")


# ws1 = wb.active
# ws1.title = "range names"

# for row in range(1, 40):
#     ws1.append(range(600))

# ws2 = wb.create_sheet(title="Pi")

# ws2['F5'] = 3.14

# ws3 = wb.create_sheet(title="Data")
# for row in range(10, 20):
#     for col in range(27, 54):
#         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

# wb.save(filename = dest_filename)

tree = ET.parse(inputFileName)
root = tree.getroot()
print(root)

# for child in root:
#     print (child.tag, child.attrib, child.text)

# filename ="result_key.txt" 
# f = open(filename,"w",encoding="utf-8")

i = 1
for child in root:
    sheet.cell(column=1, row=i+1, value=child.attrib["name"])
    i += 1
    # f.write(child.attrib["name"])
    # f.write("\n")

# f.close()


# filename ="result_source.txt" 
# f = open(filename,"w",encoding="utf-8")
i = 1
for child in root:
    sheet.cell(column=2, row=i+1, value=child.text)
    i += 1
    # f.write(child.text)
    # f.write("\n")

# f.close()


# filename ="result_translated.txt" 
# f = open(filename,"w",encoding="utf-8")

dic = {}

newroot = ET.parse(translatedFileName).getroot()

for child in newroot:
    dic[child.attrib["name"]] = child.text

i = 1
for child in root:
    if child.attrib["name"] in dic:
        sheet.cell(column=3, row=i+1, value=dic[child.attrib["name"]])
        # f.write(dic[child.attrib["name"]])
        # f.write("\n")
    else:
        sheet.cell(column=3, row=i+1, value="NOT TRANSLATED")
        # f.write("NOT TRANSLATED!!!")
        # f.write("\n")
    i += 1

wb.save(filename = dest_filename)

# f.close()














